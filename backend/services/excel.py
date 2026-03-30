"""
Excel/CSV service for import/export operations.
"""
import io
import csv
from typing import List, Dict, Tuple, Optional
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExcelError(Exception):
    """Excel operation error."""
    pass


class ProductImporter:
    """Import products from Excel/CSV."""

    REQUIRED_FIELDS = ["name", "sku", "price"]
    OPTIONAL_FIELDS = ["barcode", "category", "quantity", "cost", "description"]

    @staticmethod
    def validate_row(row: Dict, row_num: int) -> Tuple[bool, Optional[str]]:
        """Validate a single row."""
        for field in ProductImporter.REQUIRED_FIELDS:
            if not row.get(field):
                return False, f"Row {row_num}: Missing required field '{field}'"

        try:
            float(row["price"])
        except (ValueError, TypeError):
            return False, f"Row {row_num}: Invalid price format"

        if row.get("cost"):
            try:
                float(row["cost"])
            except (ValueError, TypeError):
                return False, f"Row {row_num}: Invalid cost format"

        if row.get("quantity"):
            try:
                int(row["quantity"])
            except (ValueError, TypeError):
                return False, f"Row {row_num}: Invalid quantity format"

        return True, None

    @staticmethod
    def parse_excel(file_content: bytes) -> Tuple[List[Dict], List[str]]:
        """Parse Excel file and return list of products."""
        if not HAS_OPENPYXL:
            raise ExcelError("openpyxl not installed. Install with: pip install openpyxl")

        errors = []
        products = []

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content))
            ws = wb.active

            # Get headers from first row
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(cell.value.lower().strip())

            if not headers:
                raise ExcelError("No headers found in first row")

            # Process data rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Skip empty rows
                    continue

                row_dict = {}
                for col_idx, header in enumerate(headers):
                    if col_idx < len(row):
                        row_dict[header] = row[col_idx]

                # Validate row
                is_valid, error_msg = ProductImporter.validate_row(row_dict, row_idx)
                if not is_valid:
                    errors.append(error_msg)
                    continue

                # Clean and prepare data
                product = {
                    "name": str(row_dict["name"]).strip(),
                    "sku": str(row_dict["sku"]).strip(),
                    "price": float(row_dict["price"]),
                    "barcode": row_dict.get("barcode", "").strip() if row_dict.get("barcode") else None,
                    "category": row_dict.get("category", "").strip() if row_dict.get("category") else None,
                    "quantity": int(row_dict.get("quantity", 0)) if row_dict.get("quantity") else 0,
                    "cost": float(row_dict.get("cost")) if row_dict.get("cost") else None,
                    "description": str(row_dict.get("description", "")).strip() if row_dict.get("description") else None,
                }

                products.append(product)

        except Exception as e:
            if isinstance(e, ExcelError):
                raise
            raise ExcelError(f"Failed to parse Excel file: {str(e)}")

        return products, errors

    @staticmethod
    def parse_csv(file_content: bytes) -> Tuple[List[Dict], List[str]]:
        """Parse CSV file and return list of products."""
        errors = []
        products = []

        try:
            content_str = file_content.decode('utf-8')
            reader = csv.DictReader(io.StringIO(content_str))

            if not reader.fieldnames:
                raise ExcelError("CSV file is empty")

            for row_idx, row in enumerate(reader, start=2):
                # Normalize keys
                row = {k.lower().strip(): v for k, v in row.items()}

                # Validate
                is_valid, error_msg = ProductImporter.validate_row(row, row_idx)
                if not is_valid:
                    errors.append(error_msg)
                    continue

                product = {
                    "name": str(row["name"]).strip(),
                    "sku": str(row["sku"]).strip(),
                    "price": float(row["price"]),
                    "barcode": row.get("barcode", "").strip() if row.get("barcode") else None,
                    "category": row.get("category", "").strip() if row.get("category") else None,
                    "quantity": int(row.get("quantity", 0)) if row.get("quantity") else 0,
                    "cost": float(row.get("cost")) if row.get("cost") else None,
                    "description": str(row.get("description", "")).strip() if row.get("description") else None,
                }

                products.append(product)

        except Exception as e:
            raise ExcelError(f"Failed to parse CSV file: {str(e)}")

        return products, errors


class ProductExporter:
    """Export products to Excel/CSV."""

    @staticmethod
    def to_excel(products: List[Dict], include_cost: bool = False) -> bytes:
        """Export products to Excel file."""
        if not HAS_OPENPYXL:
            raise ExcelError("openpyxl not installed")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"

        # Headers
        headers = ["ID", "Name", "SKU", "Barcode", "Category", "Quantity", "Price"]
        if include_cost:
            headers.append("Cost")
        headers.extend(["Created At", "Last Sold"])

        ws.append(headers)

        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data
        for product in products:
            row = [
                product.get("id"),
                product.get("name"),
                product.get("sku"),
                product.get("barcode"),
                product.get("category"),
                product.get("quantity"),
                product.get("price"),
            ]

            if include_cost:
                row.append(product.get("cost"))

            row.extend([
                product.get("created_at", "").isoformat() if product.get("created_at") else "",
                product.get("last_sold_at", "").isoformat() if product.get("last_sold_at") else "",
            ])

            ws.append(row)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 15
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    @staticmethod
    def to_csv(products: List[Dict], include_cost: bool = False) -> str:
        """Export products to CSV string."""
        output = io.StringIO()

        fieldnames = ["id", "name", "sku", "barcode", "category", "quantity", "price"]
        if include_cost:
            fieldnames.append("cost")
        fieldnames.extend(["created_at", "last_sold_at"])

        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()

        for product in products:
            row = {field: product.get(field, "") for field in fieldnames}
            # Convert datetime objects to ISO format
            if row.get("created_at") and hasattr(row["created_at"], "isoformat"):
                row["created_at"] = row["created_at"].isoformat()
            if row.get("last_sold_at") and hasattr(row["last_sold_at"], "isoformat"):
                row["last_sold_at"] = row["last_sold_at"].isoformat()

            writer.writerow(row)

        return output.getvalue()


class TemplateGenerator:
    """Generate import templates."""

    @staticmethod
    def generate_excel_template() -> bytes:
        """Generate an Excel template for product import."""
        if not HAS_OPENPYXL:
            raise ExcelError("openpyxl not installed")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"

        # Headers
        headers = ["Name", "SKU", "Barcode", "Category", "Quantity", "Price", "Cost", "Description"]
        ws.append(headers)

        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Add example row
        example_row = [
            "Example Product",
            "EXP001",
            "1234567890123",
            "Electronics",
            "100",
            "99.99",
            "50.00",
            "Example description",
        ]
        ws.append(example_row)

        # Set column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 10
        ws.column_dimensions["H"].width = 30

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    @staticmethod
    def generate_csv_template() -> str:
        """Generate a CSV template for product import."""
        output = io.StringIO()

        headers = ["Name", "SKU", "Barcode", "Category", "Quantity", "Price", "Cost", "Description"]
        writer = csv.writer(output)
        writer.writerow(headers)

        # Add example row
        writer.writerow([
            "Example Product",
            "EXP001",
            "1234567890123",
            "Electronics",
            "100",
            "99.99",
            "50.00",
            "Example description",
        ])

        return output.getvalue()
