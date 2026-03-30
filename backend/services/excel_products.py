"""
Шаблон xlsx (колонки A–F), как у поставщика:

  A — № п/п
  B — наименование или код партии (XS… / NXS …); строки только с кодом без цен — разделитель партии
  C — цена за ед. в ¥ («12юань», «5800 юань») → в БД только число
  D — кол-во («10шт»)
  E — «Масса, ед» в файле = доставка в ₸ («80тг»), может быть пусто
  F — сумма = цена продажи в ₸ («3000тг»), если пусто — строка пропускается

Повторяющиеся строки заголовков таблицы внутри листа пропускаются.
"""
from __future__ import annotations

import random
import re
from decimal import Decimal
from io import BytesIO
from typing import Any, Callable, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from sqlalchemy import func
from sqlalchemy.orm import Session

import models


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).replace("\u00a0", " ").strip()


def _extract_decimal(raw: Any) -> Optional[Decimal]:
    s = _cell_str(raw)
    if not s:
        return None
    t = s.lower().replace(" ", "")
    t = re.sub(r"юан[ьейя]*", "", t, flags=re.I)
    t = re.sub(r"т\.?г\.?|тенге|тг", "", t, flags=re.I)
    t = re.sub(r"шт\.?|штук[и]?", "", t, flags=re.I)
    t = t.replace(",", ".")
    m = re.search(r"(\d+(?:\.\d+)?)", t)
    if not m:
        return None
    return Decimal(m.group(1))


def _ean13_check_digit(d12: str) -> str:
    s = sum(int(d12[i]) * (3 if i % 2 else 1) for i in range(12))
    return str((10 - (s % 10)) % 10)


def _random_ean13() -> str:
    base = "".join(str(random.randint(0, 9)) for _ in range(12))
    return base + _ean13_check_digit(base)


def _unique_barcode(db: Session) -> str:
    for _ in range(80):
        cand = _random_ean13()
        exists = db.query(models.Product.id).filter(models.Product.barcode == cand).first()
        if not exists:
            return cand
    raise RuntimeError("Could not allocate unique barcode")


def _build_generated_sku(db: Session) -> str:
    last_id = db.query(func.max(models.Product.id)).scalar() or 0
    return f"AUTO-{last_id + 1:06d}"


def _compact_batch_token(name: str) -> str:
    """Код партии: убираем пробелы, верхний регистр (NXS 2502140029 → NXS2502140029)."""
    return re.sub(r"\s+", "", name.strip().upper())


def _is_batch_only_row(name: str, cny: Any, qty: Any, delivery: Any, sale: Any) -> bool:
    """Строка только с кодом партии в B, без чисел в C–F (в т.ч. без фона — как XS2502150005)."""
    if not name or len(name) > 80:
        return False
    if _cell_str(cny) or _cell_str(qty) or _cell_str(delivery) or _cell_str(sale):
        return False
    if re.search(r"[А-ЯЁа-я]", name):
        return False
    u = _compact_batch_token(name)
    if len(u) < 6 or len(u) > 40:
        return False
    return bool(re.match(r"^[A-Z0-9\-]+$", u))


def _is_inner_table_header_row(row: Tuple[Any, ...], cmap: Dict[str, int]) -> bool:
    """Повтор заголовков («Цена за ед», «Кол-во»…) внутри сгруппированного листа."""
    name_cell = _cell_str(_get(row, cmap["name"])).lower()
    if any(x in name_cell for x in ("п/п", "№п", "назван", "наимен")) and len(name_cell) < 36:
        if "цена" not in name_cell and not re.search(r"\d", name_cell):
            return True
    cny_h = _cell_str(_get(row, cmap["cny"])).lower()
    qty_h = _cell_str(_get(row, cmap["qty"])).lower()
    if "цена" in cny_h and ("ед" in cny_h or "юан" in cny_h):
        return True
    if "кол" in qty_h and len(qty_h) < 24:
        return True
    return False


def _detect_columns(header_row: Tuple[Any, ...]) -> Dict[str, int]:
    cells = [_cell_str(c).lower() for c in header_row]
    cmap: Dict[str, int] = {}
    for idx, cell in enumerate(cells):
        if any(x in cell for x in ("п/п", "№п", "№ п", "номер")) and len(cell) < 14:
            cmap.setdefault("idx", idx)
        if "цена" in cell and ("ед" in cell or "юан" in cell):
            cmap["cny"] = idx
        if "кол" in cell:
            cmap["qty"] = idx
        if "масс" in cell:
            cmap["delivery"] = idx
        if cell == "сумма" or ("сумма" in cell and "итог" not in cell):
            cmap["sale"] = idx
        if any(x in cell for x in ("назван", "наимен", "товар")) and "цена" not in cell:
            cmap["name"] = idx
    if "name" not in cmap:
        cmap["name"] = 1 if len(cells) > 1 else 0
    cmap.setdefault("idx", 0)
    cmap.setdefault("cny", 2)
    cmap.setdefault("qty", 3)
    cmap.setdefault("delivery", 4)
    cmap.setdefault("sale", 5)
    return cmap


def _find_header_row(rows: List[Tuple[Any, ...]]) -> Tuple[Optional[int], Dict[str, int]]:
    for r_idx, row in enumerate(rows[:50]):
        if not row:
            continue
        cells = [_cell_str(c).lower() for c in row if c is not None]
        joined = " ".join(cells)
        score = sum(
            1
            for kw in ("цена", "кол", "сумма", "юан", "масс", "п/п", "шт")
            if kw in joined
        )
        if score >= 3:
            return r_idx, _detect_columns(row)
    return None, {"idx": 0, "name": 1, "cny": 2, "qty": 3, "delivery": 4, "sale": 5}


def _get(row: Tuple[Any, ...], col: int) -> Any:
    if col < 0 or col >= len(row):
        return None
    return row[col]


def import_products_from_xlsx(
    db: Session,
    file_bytes: bytes,
    cny_rate: Decimal,
    progress_cb: Optional[Callable[[int, int], None]] = None,
) -> Tuple[int, List[Dict[str, Any]]]:
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return 0, [{"row": 0, "reason": "Пустой файл", "raw": None}]

    header_idx, cmap = _find_header_row(rows)
    data_start = (header_idx + 1) if header_idx is not None else 0
    skipped: List[Dict[str, Any]] = []
    created = 0
    last_batch = ""
    total_scan = max(0, len(rows) - data_start)
    if progress_cb:
        progress_cb(0, total_scan)

    for i, r_idx in enumerate(range(data_start, len(rows)), start=1):
        try:
            row = rows[r_idx]
            if not row:
                continue
            excel_line = r_idx + 1
            name_raw = _get(row, cmap["name"])
            name_s = _cell_str(name_raw)
            cny_raw = _get(row, cmap["cny"])
            qty_raw = _get(row, cmap["qty"])
            del_raw = _get(row, cmap["delivery"])
            sale_raw = _get(row, cmap["sale"])

            if _is_inner_table_header_row(row, cmap):
                continue

            if not name_s and not _cell_str(cny_raw) and not _cell_str(sale_raw):
                continue

            if _is_batch_only_row(name_s, cny_raw, qty_raw, del_raw, sale_raw):
                last_batch = name_s.strip()
                continue

            if not name_s:
                row_preview = " | ".join(_cell_str(c) for c in row[:8] if _cell_str(c))[:200] or None
                skipped.append({"row": excel_line, "reason": "Нет названия товара", "raw": row_preview})
                continue

            cny = _extract_decimal(cny_raw)
            qty_dec = _extract_decimal(qty_raw)
            delivery = _extract_decimal(del_raw)
            sale = _extract_decimal(sale_raw)

            if cny is None or cny <= 0:
                skipped.append({"row": excel_line, "reason": "Нет закупа в ¥ (цена за ед.)", "raw": _cell_str(cny_raw)})
                continue
            if sale is None or sale <= 0:
                skipped.append({"row": excel_line, "reason": "Нет суммы продажи (₸)", "raw": _cell_str(sale_raw)})
                continue
            if qty_dec is None or qty_dec <= 0:
                skipped.append({"row": excel_line, "reason": "Нет количества", "raw": _cell_str(qty_raw)})
                continue

            qty_int = int(qty_dec)
            if last_batch:
                prefix = f"[{last_batch}] "
                display_name = (prefix + name_s.strip())[:255]
            else:
                display_name = name_s[:255]

            existing_same_name = (
                db.query(models.Product.id)
                .filter(models.Product.name == display_name, models.Product.is_active.is_(True))
                .first()
            )
            if existing_same_name:
                skipped.append(
                    {
                        "row": excel_line,
                        "reason": "Уже в каталоге (то же наименование), строка пропущена",
                        "raw": display_name[:200],
                    }
                )
                continue

            purchase = cny * Decimal(str(cny_rate)) + (delivery or Decimal(0))

            sku = _build_generated_sku(db)
            while db.query(models.Product.id).filter(models.Product.sku == sku).first():
                sku = _build_generated_sku(db)

            barcode = _unique_barcode(db)

            db_product = models.Product(
                name=display_name,
                sku=sku,
                barcode=barcode,
                category=None,
                brand=None,
                description=None,
                purchase_price=purchase,
                sale_price=sale,
                cny_price=cny,
                delivery_cost_kzt=delivery,
                quantity=qty_int,
                min_quantity=0,
                is_active=True,
            )
            try:
                db.add(db_product)
                db.flush()

                db.add(
                    models.History(
                        product_id=db_product.id,
                        operation_type=models.OperationType.ADDED.value,
                        quantity_change=db_product.quantity,
                        reference_type="excel_import",
                        reference_id=db_product.id,
                        details={"message": f"Импорт Excel строка {excel_line}", "name": display_name},
                    )
                )
                db.commit()
                db.refresh(db_product)
                created += 1
            except Exception as exc:
                db.rollback()
                skipped.append(
                    {
                        "row": excel_line,
                        "reason": f"Ошибка БД: {exc!s}"[:240],
                        "raw": display_name[:200],
                    }
                )
        finally:
            if progress_cb:
                progress_cb(i, total_scan)

    return created, skipped


def export_products_xlsx(products: List[models.Product]) -> Tuple[BytesIO, str]:
    from datetime import datetime

    wb = Workbook()
    ws = wb.active
    ws.title = "Товары"
    headers = (
        "№п/п",
        "Наименование",
        "Цена за ед",
        "Кол-во",
        "Масса, ед",
        "Сумма",
    )
    ws.append(list(headers))
    for i, p in enumerate(products, start=1):
        cny = p.cny_price if p.cny_price is not None else Decimal(0)
        dlv = p.delivery_cost_kzt if p.delivery_cost_kzt is not None else Decimal(0)
        ws.append(
            [
                i,
                p.name,
                float(cny),
                int(p.quantity or 0),
                float(dlv),
                float(p.sale_price or 0),
            ]
        )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"skladpro_tovary_{datetime.utcnow().strftime('%Y-%m-%d_%H%M')}.xlsx"
    return buf, fname
